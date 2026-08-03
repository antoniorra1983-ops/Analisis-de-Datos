#!/usr/bin/env python3
"""
planilla_maniobras.py
=====================

Convierte la salida de un simulador de operación ferroviaria (CSV) en una
planilla horaria con maniobras, en formato de **varias terminales** lado a lado
(estilo "Planilla + Maniobras").

Cada viaje (tripID) se resume en una sola fila y se ubica en la columna de su
**estación de origen**. Siempre se muestran las dos terminales de cabecera
(Puerto y Limache); las terminales intermedias (El Belloto, Sargento Aldea)
aparecen solo si hay servicios que parten desde ellas.

Columnas de cada terminal:

    Viaje | Tren | Partida | N° | Inter. | Man. | Destino | M | Obs.

Maniobras (Man.), derivadas de la cadena cronológica de cada tren:

    * EV  (Entrada a Vía)  -> primer viaje del día de ese tren
    * RET (Retorno)        -> cada viaje siguiente (da vuelta en la terminal)
    * SV  (Sale de Vía)    -> fila al terminar la jornada (estaciona)

Uso básico:

    python planilla_maniobras.py Planilla_Simulador.csv -o salida.xlsx

Ver todas las opciones:

    python planilla_maniobras.py --help
"""

from __future__ import annotations

import argparse
import os
from dataclasses import dataclass, field
from datetime import datetime, time

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# --------------------------------------------------------------------------- #
# Configuración
# --------------------------------------------------------------------------- #
@dataclass
class Config:
    """Parámetros que controlan la conversión."""

    sep: str = ";"

    # Terminales de cabecera (siempre presentes). El Puerto es el extremo "sur":
    # los trenes que parten de Puerto van hacia Limache; todos los demás orígenes
    # van hacia Puerto.
    cod_puerto: str = "PUE"
    nombre_puerto: str = "Terminal Puerto"
    cod_limache: str = "LIM"
    nombre_limache: str = "Terminal Limache"

    # Terminales intermedias (puntos de retorno). Solo aparecen si hay servicios
    # que parten (o terminan) en ellas. Orden geográfico Puerto -> Limache.
    intermedios: list[tuple[str, str]] = field(
        default_factory=lambda: [("BTO", "El Belloto"), ("SGA", "Sargento Aldea")]
    )

    multiple_threshold: int = 400       # capacidad >= esto  ->  "Múltiple"
    round_minutes: bool = False         # redondear horas al minuto
    maniobras: bool = True              # derivar EV / RET / SV
    train_prefix: str = ""              # prefijo para renumerar trenes
    titulo: str = "Planilla Horaria + Maniobras — Simulador"
    fuente: str = "Arial"
    fecha_thdr: str = ""                # etiqueta de fecha en las hojas THDR (ej. "010526")

    # Nombres de columnas esperados en el CSV
    col_trip: str = "tripID"
    col_train: str = "trainID"
    col_cap: str = "trainTotalCapacity"
    col_track: str = "trackID"
    col_station: str = "stationName"
    col_arrive: str = "arriveTime"
    col_leave: str = "leaveTime"

    # Columnas de pasajeros (para las hojas de Carga de Pasajeros V1/V2).
    # Si no están en el CSV, esas hojas simplemente no se generan.
    col_arrive_pax: str = "arrivePassengers"
    col_arrive_est_pax: str = "arriveStationPassengers"
    col_leave_pax: str = "leavePassengers"
    col_leave_est_pax: str = "leaveStationPassengers"

    linea_carga: str = "Línea: Valparaíso-Limache"
    contrato_carga: str = "Todos los Contratos"


# Encabezados de cada terminal, en orden (9 columnas; la capacidad se refleja
# en la columna "M" como "Múltiple").
COLUMNAS = ["Viaje", "Tren", "Partida", "N°", "Inter.",
            "Man.", "Destino", "M", "Obs."]


# --------------------------------------------------------------------------- #
# Hojas THDR (una por vía) — formato de los archivos THDR_viaN
# --------------------------------------------------------------------------- #
# Código de estación del CSV -> nombre completo usado en el THDR.
NOMBRES_ESTACIONES = {
    "PUE": "Puerto", "BEL": "Bellavista", "FRA": "Francia", "BAR": "Baron",
    "POR": "Portales", "REC": "Recreo", "MIR": "Miramar", "VIN": "Vina del mar",
    "HOS": "Hospital", "CHO": "Chorrillos", "SLT": "El Salto", "VAL": "Valencia",
    "QUI": "Quilpue", "SOL": "El Sol", "BTO": "El Belloto", "AME": "Americas",
    "CON": "La Concepcion", "VAM": "Villa Alemana", "SGA": "Sargento Aldea",
    "PEN": "Penablanca", "LIM": "Limache",
}

# Campos de cabecera del THDR (columnas A..I).
THDR_CAMPOS = ["Viaje", "Tren", "Hora Salida Programada", "Motriz 1", "Motriz 2",
               "Unidad", "Maquinista", "Controlador", "Observación"]

THDR_COL_INICIO = 12   # primera columna de estaciones (A..I + 2 de separación)
THDR_FILA_DATOS = 6    # primera fila de datos


# --------------------------------------------------------------------------- #
# Utilidades de tiempo
# --------------------------------------------------------------------------- #
def hms_to_seconds(value: str) -> int:
    h, m, s = (int(p) for p in str(value).split(":"))
    return h * 3600 + m * 60 + s


def seconds_to_time(total: float, round_minutes: bool = False) -> time:
    total = int(round(total))
    if round_minutes:
        total = (total + 30) // 60 * 60
    h, m, s = total // 3600, (total % 3600) // 60, total % 60
    return time(h % 24, m, 0 if round_minutes else s)


def _hhmmss(total: float) -> str:
    """Hora como texto 'HH:MM:SS' a partir de segundos."""
    t = int(round(total))
    return f"{t // 3600:02d}:{(t % 3600) // 60:02d}:{t % 60:02d}"


# --------------------------------------------------------------------------- #
# 1) Resumir el CSV a una fila por viaje
# --------------------------------------------------------------------------- #
def cargar_viajes(csv_path, cfg: Config) -> pd.DataFrame:
    """Lee el CSV (ruta o buffer) y devuelve un DataFrame con una fila por tripID."""
    df = pd.read_csv(csv_path, sep=cfg.sep)

    faltan = [c for c in (cfg.col_trip, cfg.col_train, cfg.col_cap, cfg.col_track,
                          cfg.col_station, cfg.col_arrive, cfg.col_leave)
              if c not in df.columns]
    if faltan:
        raise ValueError(
            f"El CSV no tiene las columnas esperadas: {faltan}\n"
            f"Columnas encontradas: {list(df.columns)}"
        )

    viajes = []
    for trip_id, g in df.groupby(cfg.col_trip, sort=True):
        g = g.reset_index(drop=True)
        primera, ultima = g.iloc[0], g.iloc[-1]
        viajes.append({
            "trip": int(trip_id),
            "train": int(primera[cfg.col_train]),
            "cap": int(primera[cfg.col_cap]),
            "track": int(primera[cfg.col_track]),
            "orig": primera[cfg.col_station],
            "dep": primera[cfg.col_leave],
            "dep_s": hms_to_seconds(primera[cfg.col_leave]),
            "dest": ultima[cfg.col_station],
            "arr_s": hms_to_seconds(ultima[cfg.col_arrive]),
        })

    return pd.DataFrame(viajes).sort_values("dep_s").reset_index(drop=True)


# --------------------------------------------------------------------------- #
# 2) Maniobras (EV / RET) y fines de servicio (SV)
# --------------------------------------------------------------------------- #
def asignar_maniobras(viajes: pd.DataFrame) -> pd.DataFrame:
    """EV para el primer viaje de cada tren (orden cronológico), RET para el resto."""
    primer = {tren: viajes[viajes["train"] == tren].sort_values("dep_s").index[0]
              for tren in viajes["train"].unique()}
    viajes = viajes.copy()
    viajes["man"] = ["EV" if idx == primer[row.train] else "RET"
                     for idx, row in viajes.iterrows()]
    return viajes


def fines_de_servicio(viajes: pd.DataFrame) -> dict[str, list[dict]]:
    """Devuelve, por código de estación, la lista de trenes que terminan su jornada allí."""
    fines: dict[str, list[dict]] = {}
    for tren in viajes["train"].unique():
        ultimo = viajes[viajes["train"] == tren].sort_values("dep_s").iloc[-1]
        fines.setdefault(ultimo["dest"], []).append(
            {"train": tren, "cap": int(ultimo["cap"]), "end_s": ultimo["arr_s"]}
        )
    return fines


# --------------------------------------------------------------------------- #
# 3) Definir y construir las columnas (terminales)
# --------------------------------------------------------------------------- #
def columnas_visibles(viajes: pd.DataFrame, cfg: Config) -> list[dict]:
    """Lista ordenada de terminales a mostrar: Puerto, intermedias con actividad, Limache."""
    presentes = set(viajes["orig"]) | set(viajes["dest"])
    cols = [dict(code=cfg.cod_puerto, nombre=cfg.nombre_puerto, implied=cfg.cod_limache)]
    for code, nombre in cfg.intermedios:
        if code in presentes:
            cols.append(dict(code=code, nombre=nombre, implied=cfg.cod_puerto))
    cols.append(dict(code=cfg.cod_limache, nombre=cfg.nombre_limache, implied=cfg.cod_puerto))
    return cols


def _fmt_tren(num: int, cfg: Config):
    return f"{cfg.train_prefix}{num}" if cfg.train_prefix else num


def _columna_de(orig: str, track: int, cols: list[dict], cfg: Config) -> int:
    """Índice de la columna que corresponde a un viaje según su origen (con respaldo por sentido)."""
    for i, c in enumerate(cols):
        if c["code"] == orig:
            return i
    return 0 if track == 0 else len(cols) - 1  # respaldo: track 0 -> Puerto, 1 -> Limache


def construir_columna(viajes: pd.DataFrame, fines: dict, col: dict, cfg: Config) -> list[list]:
    """Filas (listas de 9 celdas) de una terminal: salidas desde col['code'] + SV que terminan allí."""
    sub = viajes[viajes["_col_code"] == col["code"]].sort_values("dep_s").reset_index(drop=True)
    filas: list[list] = []
    prev_s = None
    n = 0
    for x in sub.itertuples():
        n += 1
        inter = seconds_to_time(x.dep_s - prev_s, cfg.round_minutes) if prev_s is not None else None
        prev_s = x.dep_s
        destino = "" if x.dest == col["implied"] else x.dest
        multiple = "Múltiple" if x.cap >= cfg.multiple_threshold else ""
        maniobra = x.man if cfg.maniobras else ""
        filas.append([x.trip, _fmt_tren(x.train, cfg),
                      seconds_to_time(x.dep_s, cfg.round_minutes),
                      n, inter, maniobra, destino, multiple, ""])

    if cfg.maniobras:
        for r in sorted(fines.get(col["code"], []), key=lambda d: d["end_s"]):
            multiple = "Múltiple" if r["cap"] >= cfg.multiple_threshold else ""
            filas.append(["", _fmt_tren(r["train"], cfg), None, "", None,
                          "SV", "", multiple, "Estaciona"])
    return filas


def construir_tablas(viajes: pd.DataFrame, cfg: Config) -> tuple[list[dict], list[list[list]]]:
    """Devuelve (columnas, tablas) en paralelo: una lista de filas por terminal."""
    viajes = asignar_maniobras(viajes) if cfg.maniobras else viajes.assign(man="")
    fines = fines_de_servicio(viajes) if cfg.maniobras else {}
    cols = columnas_visibles(viajes, cfg)

    # asignar cada viaje a su columna por origen
    viajes = viajes.copy()
    viajes["_col_code"] = [cols[_columna_de(r.orig, r.track, cols, cfg)]["code"]
                           for r in viajes.itertuples()]

    tablas = [construir_columna(viajes, fines, c, cfg) for c in cols]
    return cols, tablas


# --------------------------------------------------------------------------- #
# 4) Escribir el Excel con formato
# --------------------------------------------------------------------------- #
def cargar_paradas(csv_path, cfg: Config) -> list[dict]:
    """Lee el CSV y devuelve, por viaje, el detalle de cada parada.

    Cada elemento: {trip, train, cap, track, paradas: [{est, lleg, sal}, ...]}
    Es el detalle que necesitan las hojas THDR (llegada/salida por estación).
    """
    df = pd.read_csv(csv_path, sep=cfg.sep)
    faltan = [c for c in (cfg.col_trip, cfg.col_train, cfg.col_cap, cfg.col_track,
                          cfg.col_station, cfg.col_arrive, cfg.col_leave)
              if c not in df.columns]
    if faltan:
        raise ValueError(f"El CSV no tiene las columnas esperadas: {faltan}")

    viajes = []
    tiene_pax = all(c in df.columns for c in
                    (cfg.col_arrive_pax, cfg.col_arrive_est_pax,
                     cfg.col_leave_pax, cfg.col_leave_est_pax))
    for trip_id, g in df.groupby(cfg.col_trip, sort=True):
        g = g.reset_index(drop=True)
        paradas = []
        for _, r in g.iterrows():
            p = {"est": str(r[cfg.col_station]).strip(),
                 "lleg": hms_to_seconds(r[cfg.col_arrive]),
                 "sal": hms_to_seconds(r[cfg.col_leave])}
            if tiene_pax:
                # carga = pasajeros a bordo al salir; suben = los que abordan aquí
                p["carga"] = int(r[cfg.col_leave_pax])
                p["suben"] = int(r[cfg.col_arrive_est_pax]) - int(r[cfg.col_leave_est_pax])
            paradas.append(p)
        viajes.append({
            "trip": int(trip_id),
            "train": int(g.iloc[0][cfg.col_train]),
            "cap": int(g.iloc[0][cfg.col_cap]),
            "track": int(g.iloc[0][cfg.col_track]),
            "pax": tiene_pax,
            "paradas": paradas,
        })
    return viajes


def _orden_estaciones(paradas_viajes: list[dict], track: int) -> list[str]:
    """Orden de estaciones de una vía: el recorrido más largo de esa vía."""
    recorridos = [[p["est"] for p in v["paradas"]]
                  for v in paradas_viajes if v["track"] == track]
    if not recorridos:
        return []
    base = max(recorridos, key=len)
    # Completar con estaciones que aparezcan en otros recorridos de la misma vía
    for rec in recorridos:
        for i, est in enumerate(rec):
            if est not in base:
                anterior = rec[i - 1] if i else None
                pos = base.index(anterior) + 1 if anterior in base else len(base)
                base.insert(pos, est)
    return base


def agregar_hojas_thdr(wb: Workbook, paradas_viajes: list[dict], cfg: Config) -> list[str]:
    """Añade al libro una hoja THDR por vía (llegada/salida por estación).

    Vía 1 = trenes que van de Puerto a Limache (track 0);
    Vía 2 = sentido contrario (track 1).
    """
    NAVY = "1F3864"
    thin = Side(style="thin", color="B0B0B0")
    borde = Border(left=thin, right=thin, top=thin, bottom=thin)
    f_cell = Font(name=cfg.fuente, size=9)
    f_hdr = Font(name=cfg.fuente, bold=True, color="FFFFFF", size=9)
    f_est = Font(name=cfg.fuente, bold=True, size=10, color=NAVY)
    center = Alignment(horizontal="center", vertical="center")
    wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
    fmt_time = "h:mm" if cfg.round_minutes else "h:mm:ss"

    creadas = []
    for via, track in ((1, 0), (2, 1)):
        orden = _orden_estaciones(paradas_viajes, track)
        if not orden:
            continue
        ws = wb.create_sheet(f"THDR Vía {via}")

        # Fila 1: fecha (si se indicó) y nombres de estación
        if cfg.fecha_thdr:
            c = ws.cell(1, 1, cfg.fecha_thdr)
            c.font = Font(name=cfg.fuente, bold=True, size=10)
        col_de_est = {}
        for i, est in enumerate(orden):
            base = THDR_COL_INICIO + i * 2
            col_de_est[est] = base
            c = ws.cell(1, base, NOMBRES_ESTACIONES.get(est, est))
            c.font, c.alignment = f_est, center

        # Fila 2: encabezados
        for j, campo in enumerate(THDR_CAMPOS):
            c = ws.cell(2, 1 + j, campo)
            c.font, c.alignment, c.border = f_hdr, wrap, borde
            c.fill = PatternFill("solid", fgColor=NAVY)
        ultimo = orden[-1]
        for est in orden:
            base = col_de_est[est]
            etiquetas = ["Hora Llegada"] if est == ultimo else ["Hora Llegada", "Hora Salida"]
            for k, etq in enumerate(etiquetas):
                c = ws.cell(2, base + k, etq)
                c.font, c.alignment, c.border = f_hdr, wrap, borde
                c.fill = PatternFill("solid", fgColor=NAVY)

        # Datos (ordenados por hora de salida del origen)
        viajes_via = sorted((v for v in paradas_viajes if v["track"] == track),
                            key=lambda v: v["paradas"][0]["sal"])
        for i, v in enumerate(viajes_via):
            r = THDR_FILA_DATOS + i
            ws.cell(r, 1, v["trip"]).alignment = center
            ws.cell(r, 2, v["train"]).alignment = center
            prog = ws.cell(r, 3, seconds_to_time(v["paradas"][0]["sal"], True))
            prog.number_format, prog.alignment = "h:mm", center
            if v["cap"] >= cfg.multiple_threshold:
                ws.cell(r, 6, "M").alignment = center     # Unidad: M = múltiple
            for j in range(1, len(THDR_CAMPOS) + 1):
                cc = ws.cell(r, j)
                cc.font, cc.border = f_cell, borde

            primera, ultima_p = v["paradas"][0], v["paradas"][-1]
            for p in v["paradas"]:
                base = col_de_est.get(p["est"])
                if base is None:
                    continue
                # El origen no tiene llegada; el destino final no tiene salida.
                if p is not primera:
                    c = ws.cell(r, base, seconds_to_time(p["lleg"], cfg.round_minutes))
                    c.number_format, c.alignment, c.font, c.border = fmt_time, center, f_cell, borde
                if p is not ultima_p and p["est"] != ultimo:
                    c = ws.cell(r, base + 1, seconds_to_time(p["sal"], cfg.round_minutes))
                    c.number_format, c.alignment, c.font, c.border = fmt_time, center, f_cell, borde

        # Anchos y vista
        for j, w in enumerate([6, 6, 11, 8, 8, 7, 26, 14, 14]):
            ws.column_dimensions[get_column_letter(1 + j)].width = w
        ws.column_dimensions[get_column_letter(10)].width = 2
        ws.column_dimensions[get_column_letter(11)].width = 2
        for est in orden:
            base = col_de_est[est]
            for k in range(2 if est != ultimo else 1):
                ws.column_dimensions[get_column_letter(base + k)].width = 10
        ws.freeze_panes = ws.cell(THDR_FILA_DATOS, 3)
        ws.sheet_view.showGridLines = False
        creadas.append(ws.title)
    return creadas


# --------------------------------------------------------------------------- #
# Hojas V1 / V2 — Carga de Pasajeros (formato EXPORT Carga Pasajeros)
# --------------------------------------------------------------------------- #
CARGA_CAMPOS = ["N° THDR", "N° Viaje", "Tren", "Fecha", "Hora Origen", "Hora Fin",
                "Motriz 1", "Motriz 2"]
CARGA_FINALES = ["Total a Bordo", "Carga Máxima", "Estación Máxima"]
CARGA_TITULO = "CARGA DE PASAJEROS POR SERVICIO para Filial EFE-VALPO"
CARGA_FILA_ENC = 10     # fila de encabezados
CARGA_FILA_DATOS = 11   # primera fila de datos


def agregar_hojas_carga(wb: Workbook, paradas_viajes: list[dict], cfg: Config) -> list[str]:
    """Añade las hojas V1 y V2 con la carga de pasajeros por servicio.

    Replica el formato de los archivos *EXPORT Carga Pasajeros*: una fila por
    viaje, con los pasajeros a bordo al salir de cada estación, más el total
    transportado, la carga máxima y en qué estación se alcanza.
    """
    if not any(v.get("pax") for v in paradas_viajes):
        return []                      # el CSV no trae columnas de pasajeros

    NAVY = "1F3864"
    thin = Side(style="thin", color="B0B0B0")
    borde = Border(left=thin, right=thin, top=thin, bottom=thin)
    f_cell = Font(name=cfg.fuente, size=9)
    f_hdr = Font(name=cfg.fuente, bold=True, color="FFFFFF", size=9)
    f_bold = Font(name=cfg.fuente, bold=True, size=9)
    center = Alignment(horizontal="center", vertical="center")
    wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
    creado_el = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

    creadas = []
    for via, track in ((1, 0), (2, 1)):
        orden = _orden_estaciones(paradas_viajes, track)
        if not orden:
            continue
        ws = wb.create_sheet(f"V{via}")

        # Cabecera del reporte
        ws.cell(1, 1, CARGA_TITULO).font = Font(name=cfg.fuente, bold=True, size=11, color=NAVY)
        meta = [(3, "Fecha creación archivo ", creado_el),
                (4, "Desde:", cfg.fecha_thdr), (5, "Hasta:", cfg.fecha_thdr),
                (6, "Línea:", cfg.linea_carga), (7, "Vía:", f"Vía {via}"),
                (8, "Contrato:", cfg.contrato_carga)]
        for fila, etq, val in meta:
            ws.cell(fila, 1, etq).font = f_bold
            ws.cell(fila, 2, val).font = f_cell

        # Encabezados: campos + estaciones (códigos) + totales
        encabezados = CARGA_CAMPOS + orden + CARGA_FINALES
        for j, h in enumerate(encabezados):
            c = ws.cell(CARGA_FILA_ENC, 1 + j, h)
            c.font, c.alignment, c.border = f_hdr, wrap, borde
            c.fill = PatternFill("solid", fgColor=NAVY)

        col_de_est = {est: len(CARGA_CAMPOS) + 1 + i for i, est in enumerate(orden)}
        col_total = len(CARGA_CAMPOS) + len(orden) + 1

        viajes_via = sorted((v for v in paradas_viajes if v["track"] == track),
                            key=lambda v: v["paradas"][0]["sal"])
        for i, v in enumerate(viajes_via):
            r = CARGA_FILA_DATOS + i
            primera, ultima = v["paradas"][0], v["paradas"][-1]
            ws.cell(r, 1, v["trip"])                                  # N° THDR
            ws.cell(r, 2, "")                                         # N° Viaje (no viene en el CSV)
            ws.cell(r, 3, v["train"])
            ws.cell(r, 4, cfg.fecha_thdr)
            ws.cell(r, 5, _hhmmss(primera["sal"]))                    # Hora Origen (texto)
            ws.cell(r, 6, _hhmmss(ultima["lleg"]))                    # Hora Fin (texto)
            ws.cell(r, 7, "")                                         # Motriz 1
            ws.cell(r, 8, "")                                         # Motriz 2

            cargas = {}
            for p in v["paradas"]:
                col = col_de_est.get(p["est"])
                if col is not None:
                    ws.cell(r, col, p.get("carga", 0))
                    cargas[p["est"]] = p.get("carga", 0)
            maxima = max(cargas.values()) if cargas else 0
            ws.cell(r, col_total, sum(p.get("suben", 0) for p in v["paradas"]))
            ws.cell(r, col_total + 1, maxima)
            ws.cell(r, col_total + 2,
                    " - ".join(e for e in orden if cargas.get(e) == maxima) if maxima else "")

            for j in range(1, col_total + 3):
                cc = ws.cell(r, j)
                cc.font, cc.border = f_cell, borde
                cc.alignment = center if j != col_total + 2 else Alignment(
                    horizontal="left", vertical="center")

        # Anchos y vista
        for j, w in enumerate([9, 10, 7, 11, 11, 10, 9, 9]):
            ws.column_dimensions[get_column_letter(1 + j)].width = w
        for est in orden:
            ws.column_dimensions[get_column_letter(col_de_est[est])].width = 6
        ws.column_dimensions[get_column_letter(col_total)].width = 12
        ws.column_dimensions[get_column_letter(col_total + 1)].width = 12
        ws.column_dimensions[get_column_letter(col_total + 2)].width = 24
        ws.freeze_panes = ws.cell(CARGA_FILA_DATOS, 4)
        ws.sheet_view.showGridLines = False
        creadas.append(ws.title)
    return creadas


def construir_workbook(cols: list[dict], tablas: list[list[list]],
                       cfg: Config, hora_inicio: str, origen_archivo: str,
                       paradas_viajes: list[dict] | None = None) -> Workbook:
    """Arma el libro Excel (en memoria) con N terminales lado a lado.

    Si se pasa `paradas_viajes` (de `cargar_paradas`), añade además una hoja
    THDR por vía, quedando 3 hojas en total.
    """
    NAVY, GRIS, VERDE = "1F3864", "F2F2F2", "E2EFDA"
    NCOL = len(COLUMNAS)            # 9 columnas por bloque
    PASO = NCOL + 1                 # + 1 columna separadora
    total_cols = PASO * len(cols) - 1

    thin = Side(style="thin", color="B0B0B0")
    borde = Border(left=thin, right=thin, top=thin, bottom=thin)
    f_cell = Font(name=cfg.fuente, size=9)
    f_hdr = Font(name=cfg.fuente, bold=True, color="FFFFFF", size=9)
    f_term = Font(name=cfg.fuente, bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal="center", vertical="center")
    left_al = Alignment(horizontal="left", vertical="center")
    fmt_time = "h:mm" if cfg.round_minutes else "h:mm:ss"

    wb = Workbook()
    ws = wb.active
    ws.title = "Planilla + Maniobras"

    # Título
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    c = ws.cell(1, 1, cfg.titulo)
    c.font = Font(name=cfg.fuente, bold=True, size=13, color=NAVY)
    c.alignment = left_al

    # Metadatos
    ws.cell(3, 1, "Inicio").font = Font(name=cfg.fuente, bold=True, size=9)
    hi = ws.cell(3, 2, seconds_to_time(hms_to_seconds(hora_inicio), True))
    hi.font, hi.number_format = f_cell, "h:mm"
    ws.cell(3, 6, "Origen:").font = Font(name=cfg.fuente, bold=True, size=9)
    ws.cell(3, 7, origen_archivo).font = f_cell
    ws.cell(4, 1, "Versión").font = Font(name=cfg.fuente, bold=True, size=9)
    ws.cell(4, 2, "Simulador").font = f_cell

    bases = [1 + i * PASO for i in range(len(cols))]

    # Rótulos de terminal
    for base, col in zip(bases, cols):
        ws.merge_cells(start_row=5, start_column=base, end_row=5, end_column=base + NCOL - 1)
        t = ws.cell(5, base, col["nombre"])
        t.font, t.alignment = f_term, center
        t.fill = PatternFill("solid", fgColor=NAVY)

    # Encabezados
    for base in bases:
        for j, h in enumerate(COLUMNAS):
            cc = ws.cell(6, base + j, h)
            cc.font, cc.alignment, cc.border = f_hdr, center, borde
            cc.fill = PatternFill("solid", fgColor=NAVY)

    # Datos
    def volcar(filas: list[list], base: int) -> None:
        for i, fila in enumerate(filas):
            r = 7 + i
            for j, val in enumerate(fila):
                cc = ws.cell(r, base + j, val if val != "" else None)
                cc.border, cc.font = borde, f_cell
                if j in (2, 4):                       # Partida, Inter. -> hora
                    cc.number_format = fmt_time
                    cc.alignment = center
                elif j in (0, 1, 3, 5, 6, 7):
                    cc.alignment = center
                else:
                    cc.alignment = left_al
            man = fila[5]
            relleno = VERDE if man == "EV" else GRIS if man == "SV" else None
            if relleno:
                for j in range(NCOL):
                    ws.cell(r, base + j).fill = PatternFill("solid", fgColor=relleno)
            ws.cell(r, base + 5).font = Font(
                name=cfg.fuente, size=9, bold=True,
                color="375623" if man == "EV" else "7F7F7F" if man == "SV" else "000000",
            )

    for base, filas in zip(bases, tablas):
        volcar(filas, base)

    # Anchos / vista
    anchos = [6, 6, 9, 5, 8, 6, 7, 9, 15]
    for base in bases:
        for j, w in enumerate(anchos):
            ws.column_dimensions[get_column_letter(base + j)].width = w
        if base + NCOL <= total_cols:
            ws.column_dimensions[get_column_letter(base + NCOL)].width = 2  # separador
    ws.freeze_panes = "A7"
    ws.sheet_view.showGridLines = False

    # Hojas THDR (una por vía) y hojas V1/V2 de carga de pasajeros
    if paradas_viajes:
        agregar_hojas_thdr(wb, paradas_viajes, cfg)
        agregar_hojas_carga(wb, paradas_viajes, cfg)
    return wb


def escribir_excel(cols, tablas, out_path: str, cfg: Config,
                   hora_inicio: str, origen_archivo: str,
                   paradas_viajes: list[dict] | None = None) -> None:
    wb = construir_workbook(cols, tablas, cfg, hora_inicio, origen_archivo, paradas_viajes)
    os.makedirs(os.path.dirname(os.path.abspath(out_path)), exist_ok=True)
    wb.save(out_path)


# --------------------------------------------------------------------------- #
# Orquestador
# --------------------------------------------------------------------------- #
def convertir(csv_path: str, out_path: str, cfg: Config) -> dict:
    """Pipeline completo CSV -> XLSX en disco. Devuelve un resumen."""
    viajes = cargar_viajes(csv_path, cfg)
    cols, tablas = construir_tablas(viajes, cfg)
    hora_inicio = viajes["dep"].iloc[0]
    paradas_viajes = cargar_paradas(csv_path, cfg)
    escribir_excel(cols, tablas, out_path, cfg, hora_inicio,
                   os.path.basename(csv_path), paradas_viajes)

    return {
        "viajes_total": len(viajes),
        "trenes": int(viajes["train"].nunique()),
        "terminales": {c["nombre"]: len(t) for c, t in zip(cols, tablas)},
        "thdr": {f"Vía {v}": sum(1 for x in paradas_viajes if x["track"] == t)
                 for v, t in ((1, 0), (2, 1))},
        "salida": out_path,
    }


# =========================================================================== #
# CONVERSIÓN INVERSA
# Planilla + Maniobras (.xls)  ->  entrada del simulador (.xls, estilo Planillaprueba2)
#
# Formato de salida (9 columnas, una fila por servicio, ordenadas por hora):
#   hora | origen | unidades | destino | unidades | destino | "servicio" | tren | 406
# (xlrd y xlwt se importan de forma diferida para no exigirlos en el flujo CSV.)
# =========================================================================== #

# Etiqueta de terminal -> código de estación
_TERMINALES_ETIQUETA = [("puerto", "PUE"), ("belloto", "BTO"),
                        ("sargento", "SGA"), ("aldea", "SGA"), ("limache", "LIM")]
# Códigos numéricos (zonas) de la columna Destino: 6 = Limache, 4 = Sargento Aldea.
ZONA_A_ESTACION = {6: "LIM", 4: "SGA", 5: "BTO"}
_ESTACIONES = {"PUE", "LIM", "SGA", "BTO", "AME", "PEN"}
_ENCABEZADOS_PM = ["Viaje", "Tren", "Partida", "N°", "Inter.", "Man.", "Destino", "M", "Obs.", "Capacidad"]


def _pm_norm(s) -> str:
    return str(s).strip().lower()


def _pm_entero_pos(v) -> bool:
    return isinstance(v, (int, float)) and float(v) == int(v) and int(v) > 0


def _pm_val(sh, r, c):
    if c is None:
        return ""
    v = sh.cell_value(r, c)
    return int(v) if isinstance(v, float) and v == int(v) else v


def _pm_hora_seg(sh, wb, r, c):
    import xlrd
    if c is None:
        return None
    if sh.cell_type(r, c) == xlrd.XL_CELL_DATE:
        t = xlrd.xldate_as_tuple(sh.cell_value(r, c), wb.datemode)
        return t[3] * 3600 + t[4] * 60 + t[5]
    v = sh.cell_value(r, c)
    if isinstance(v, str) and ":" in v:
        p = [int(x) for x in v.split(":")]
        return p[0] * 3600 + p[1] * 60 + (p[2] if len(p) > 2 else 0)
    return None


def _pm_detectar(sh):
    """Localiza la fila de encabezados y los bloques (terminal -> columnas)."""
    fila_rot = fila_enc = None
    for r in range(min(15, sh.nrows)):
        textos = [_pm_norm(sh.cell_value(r, c)) for c in range(sh.ncols)]
        if any("viaje" in t for t in textos) and any("partida" in t for t in textos):
            fila_enc = r
        if any("terminal" in t or "belloto" in t or "sargento" in t or "aldea" in t for t in textos):
            fila_rot = r
    if fila_enc is None:
        raise ValueError("No se encontró la fila de encabezados (con 'Viaje'/'Partida').")
    if fila_rot is None:
        fila_rot = fila_enc - 1

    inicios = []
    for c in range(sh.ncols):
        etq = _pm_norm(sh.cell_value(fila_rot, c))
        if not etq:
            continue
        code = next((cod for clave, cod in _TERMINALES_ETIQUETA if clave in etq), None)
        if code:
            inicios.append((c, code))

    bloques = []
    for i, (c0, code) in enumerate(inicios):
        c1 = inicios[i + 1][0] if i + 1 < len(inicios) else sh.ncols
        colmap = {}
        for c in range(c0, c1):
            h = _pm_norm(sh.cell_value(fila_enc, c))
            for nombre in _ENCABEZADOS_PM:
                if _pm_norm(nombre) == h or (nombre == "N°" and h in ("n°", "n")):
                    colmap[nombre] = c
        bloques.append((code, colmap))
    return fila_enc, bloques


def _pm_destino(origen: str, via, raw) -> str:
    """Estación destino del viaje. Para salidas de Puerto se deduce de la vía
    (6→LIM, 4→SGA); el resto de terminales van a Puerto. Un código de estación
    explícito en la celda (p. ej. sábado) tiene prioridad."""
    if isinstance(raw, str) and raw.strip().upper() in _ESTACIONES:
        return raw.strip().upper()
    if origen == "PUE":
        try:
            return ZONA_A_ESTACION.get(int(via), "LIM")
        except (TypeError, ValueError):
            return "LIM"
    return "PUE"


def _pm_a_entero(v):
    """Convierte a int un valor que puede venir como número o como texto ('6')."""
    try:
        if isinstance(v, str):
            v = v.strip()
            if v == "":
                return None
        return int(float(v))
    except (TypeError, ValueError):
        return None


def _pm_mapas_por_tren(sh, fila_enc, bloques):
    """Recorre toda la hoja y arma dos mapas por tren:
       - via_por_tren: el valor (fijo) de la columna Destino de ese tren -> es la vía.
       - trenes_multiple: trenes con 'Múltiple' en CUALQUIER columna de alguna de sus
         filas (en estos archivos la marca suele ir en la columna Obs., no en M)."""
    via_por_tren: dict[int, int] = {}
    trenes_multiple: set[int] = set()
    for _code, colmap in bloques:
        cols = list(colmap.values())
        c_ini, c_fin = min(cols), max(cols)
        for r in range(fila_enc + 1, sh.nrows):
            tren = _pm_a_entero(_pm_val(sh, r, colmap.get("Tren")))
            if tren is None or tren <= 0:
                continue
            if any("ltiple" in str(_pm_val(sh, r, c)).lower() for c in range(c_ini, c_fin + 1)):
                trenes_multiple.add(tren)
            via = _pm_a_entero(_pm_val(sh, r, colmap.get("Destino")))
            if tren not in via_por_tren and via is not None:
                via_por_tren[tren] = via
    return via_por_tren, trenes_multiple


def _pm_abrir(origen):
    """Abre un .xls desde ruta o bytes y devuelve el workbook (xlrd)."""
    import xlrd
    if isinstance(origen, (bytes, bytearray)):
        return xlrd.open_workbook(file_contents=bytes(origen))
    return xlrd.open_workbook(origen)


def _pm_contar_salidas(sh, wb, fila_enc, bloques) -> int:
    """Cuenta filas que son salidas reales (con N° de viaje y hora)."""
    n = 0
    for _code, colmap in bloques:
        for r in range(fila_enc + 1, sh.nrows):
            if _pm_entero_pos(_pm_val(sh, r, colmap.get("Viaje"))) and \
               _pm_hora_seg(sh, wb, r, colmap.get("Partida")) is not None:
                n += 1
    return n


def _elegir_hoja(wb, hoja_preferida=None) -> str:
    """Nombre de la hoja con formato Planilla + Maniobras (sin depender del nombre).

    Prioridad: nombre pedido si existe -> hoja con estructura válida y MÁS salidas
    (desempata el nombre que contenga 'maniobra') -> heurística de nombre -> 1ª hoja.
    """
    nombres = wb.sheet_names()
    if hoja_preferida and hoja_preferida in nombres:
        return hoja_preferida
    candidatas = []  # (0 si nombre tiene 'maniobra' si no 1, -salidas, nombre)
    for n in nombres:
        sh = wb.sheet_by_name(n)
        try:
            fila_enc, bloques = _pm_detectar(sh)
        except Exception:
            continue
        ndep = _pm_contar_salidas(sh, wb, fila_enc, bloques)
        if ndep > 0:
            candidatas.append((0 if "maniobra" in n.lower() else 1, -ndep, n))
    if candidatas:
        candidatas.sort()
        return candidatas[0][2]
    for clave in ("maniobra", "planilla"):
        for n in nombres:
            if clave in n.lower():
                return n
    return nombres[0]


def elegir_hoja_pm(origen, hoja_preferida=None) -> str:
    """Devuelve el nombre de la hoja Planilla + Maniobras de un archivo (ruta o bytes)."""
    return _elegir_hoja(_pm_abrir(origen), hoja_preferida)


def listar_hojas(origen) -> list[str]:
    """Lista los nombres de hoja de un archivo .xls (ruta o bytes)."""
    return _pm_abrir(origen).sheet_names()


def leer_planilla_maniobras(origen, hoja=None, via_defecto: int = 1) -> list[dict]:
    """Extrae los servicios de una hoja Planilla + Maniobras. `origen` puede ser ruta o bytes.

    Si `hoja` es None (o no existe en el archivo), la hoja correcta se detecta
    automáticamente por su estructura, sin importar cómo se llame.

    La vía (columnas C/E del simulador) sale de la columna Destino, que es fija
    por tren, y se arrastra a todos sus viajes (incluida la vuelta a Puerto).
    Un tren es doble si aparece 'Múltiple' en alguna de sus filas.
    """
    wb = _pm_abrir(origen)
    sh = wb.sheet_by_name(_elegir_hoja(wb, hoja))
    fila_enc, bloques = _pm_detectar(sh)
    via_por_tren, trenes_multiple = _pm_mapas_por_tren(sh, fila_enc, bloques)

    salidas = []
    for code, colmap in bloques:
        for r in range(fila_enc + 1, sh.nrows):
            viaje = _pm_val(sh, r, colmap.get("Viaje"))
            hora = _pm_hora_seg(sh, wb, r, colmap.get("Partida"))
            tren = _pm_val(sh, r, colmap.get("Tren"))
            if not _pm_entero_pos(viaje) or hora is None or not _pm_entero_pos(tren):
                continue  # filas de paso (sin viaje) o SV (sin hora) se omiten
            tren = int(tren)
            via = via_por_tren.get(tren, via_defecto)
            salidas.append({
                "hora": hora, "origen": code, "via": via,
                "destino": _pm_destino(code, via, _pm_val(sh, r, colmap.get("Destino"))),
                "tren": tren, "unidades": 2 if tren in trenes_multiple else 1,
            })
    salidas.sort(key=lambda d: (d["hora"], d["origen"]))
    return salidas


def escribir_simulador_xls(salidas: list[dict], destino, constante: int = 406) -> None:
    """Escribe el formato plano del simulador como .xls. `destino` puede ser ruta o un buffer.

    `constante` es la capacidad de pasajeros POR UNIDAD (406). La última columna
    guarda la capacidad total = constante * unidades (406 simple, 812 doble).

    Formato de celdas:
      * Columna A (hora): TEXTO con la forma "HH:MM:SS".
      * Columnas numéricas (vía, tren, capacidad): formato número.
      * Resto (origen, destino, "servicio"): texto.
    """
    import xlwt
    wb = xlwt.Workbook(encoding="utf-8")
    ws = wb.add_sheet("Hoja1")
    estilo_texto = xlwt.easyxf(num_format_str="@")    # formato texto
    estilo_numero = xlwt.easyxf(num_format_str="0")   # formato número (entero)
    for i, s in enumerate(salidas):
        h = int(s["hora"])
        hora_txt = f"{h // 3600:02d}:{(h % 3600) // 60:02d}:{h % 60:02d}"
        ws.write(i, 0, hora_txt, estilo_texto)                       # A · hora (texto HH:MM:SS)
        ws.write(i, 1, s["origen"])                                  # B · texto
        ws.write(i, 2, s["via"], estilo_numero)                      # C · vía (número)
        ws.write(i, 3, s["destino"])                                 # D · texto
        ws.write(i, 4, s["via"], estilo_numero)                      # E · vía (número)
        ws.write(i, 5, s["destino"])                                 # F · texto
        ws.write(i, 6, "servicio")                                   # G · texto
        ws.write(i, 7, s["tren"], estilo_numero)                     # H · tren (número)
        ws.write(i, 8, constante * s["unidades"], estilo_numero)     # I · capacidad (número)
    wb.save(destino)


def simulador_a_bytes(salidas: list[dict], constante: int = 406) -> bytes:
    """Devuelve el .xls del simulador como bytes (para descargar en la web)."""
    import io
    buf = io.BytesIO()
    escribir_simulador_xls(salidas, buf, constante)
    return buf.getvalue()


def convertir_a_simulador(entrada_xls: str, salida_xls: str,
                          hoja=None, constante: int = 406) -> dict:
    """Pipeline: Planilla + Maniobras (.xls) -> entrada del simulador (.xls).

    Si `hoja` es None, se detecta automáticamente la hoja con formato
    Planilla + Maniobras (sin importar su nombre).
    """
    from collections import Counter
    hoja_usada = elegir_hoja_pm(entrada_xls, hoja)
    salidas = leer_planilla_maniobras(entrada_xls, hoja_usada)
    escribir_simulador_xls(salidas, salida_xls, constante)
    return {
        "servicios": len(salidas),
        "por_origen": dict(Counter(s["origen"] for s in salidas)),
        "hoja": hoja_usada,
        "salida": salida_xls,
    }


# --------------------------------------------------------------------------- #
# CLI
# --------------------------------------------------------------------------- #
def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="Convierte el CSV del simulador a una planilla horaria con "
                    "maniobras en formato de varias terminales (.xlsx).",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    p.add_argument("input", help="Ruta del CSV del simulador.")
    p.add_argument("-o", "--output", default="Planilla_Maniobras.xlsx",
                   help="Ruta del archivo .xlsx de salida.")
    p.add_argument("--sep", default=";", help="Separador de columnas del CSV.")

    g = p.add_argument_group("Terminales")
    g.add_argument("--cod-puerto", default="PUE", help="Código de la Terminal Puerto.")
    g.add_argument("--cod-limache", default="LIM", help="Código de la Terminal Limache.")
    g.add_argument("--nombre-puerto", default="Terminal Puerto")
    g.add_argument("--nombre-limache", default="Terminal Limache")

    o = p.add_argument_group("Opciones de formato")
    o.add_argument("--multiple-threshold", type=int, default=400,
                   help="Capacidad mínima para marcar el tren como 'Múltiple'.")
    o.add_argument("--round-minutes", action="store_true",
                   help="Redondear las horas al minuto (descarta los segundos).")
    o.add_argument("--no-maniobras", action="store_true",
                   help="No derivar EV/RET/SV (genera solo la planilla básica).")
    o.add_argument("--train-prefix", default="",
                   help="Prefijo para renumerar los trenes (ej. '60').")
    o.add_argument("--titulo", default="Planilla Horaria + Maniobras — Simulador")
    o.add_argument("--fuente", default="Arial")
    return p


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    cfg = Config(
        sep=args.sep,
        cod_puerto=args.cod_puerto, cod_limache=args.cod_limache,
        nombre_puerto=args.nombre_puerto, nombre_limache=args.nombre_limache,
        multiple_threshold=args.multiple_threshold,
        round_minutes=args.round_minutes, maniobras=not args.no_maniobras,
        train_prefix=args.train_prefix, titulo=args.titulo, fuente=args.fuente,
    )
    resumen = convertir(args.input, args.output, cfg)
    print("Planilla generada correctamente.")
    print(f"  Viajes procesados : {resumen['viajes_total']}")
    print(f"  Trenes            : {resumen['trenes']}")
    for nombre, n in resumen["terminales"].items():
        print(f"  {nombre:<18}: {n} filas")
    print(f"  Archivo           : {resumen['salida']}")
    return 0


def _corriendo_en_streamlit() -> bool:
    try:
        from streamlit.runtime import exists
        return exists()
    except Exception:
        try:
            from streamlit.runtime.scriptrunner import get_script_run_ctx
            return get_script_run_ctx() is not None
        except Exception:
            return False


if __name__ == "__main__":
    if _corriendo_en_streamlit():
        import streamlit_app  # noqa: F401  -> renderiza la app si apuntan aquí por error
    else:
        raise SystemExit(main())
